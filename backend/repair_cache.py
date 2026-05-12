#!/usr/bin/env python3
import hashlib
import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).parent
EXPLANATION_CACHE = BACKEND_DIR / "explanation_cache.xlsx"
RECIPE_CACHE = BACKEND_DIR / "recipe_cache.xlsx"


def clean_recipe_text(text):
    """Strip **bold**, *italic*, and collapse whitespace. Normalize for consistent hashing."""
    if not text:
        return ""
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'(?:^|(?<=\s))\*([^*/\s]+)\*(?=\s|$)', r'\1', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def clean_dish_name(text):
    """Strip markdown from dish name, collapse whitespace, preserve case for hashing."""
    if not text:
        return ""
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'(?:^|(?<=\s))\*([^*/\s]+)\*(?=\s|$)', r'\1', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def extract_dish_name_from_recipe(recipe_text):
    """Try to extract dish name from recipe content if dish_name is corrupted."""
    if not recipe_text:
        return "unknown_recipe"

    first_line = recipe_text.split('\n')[0].strip()
    first_line_clean = clean_dish_name(first_line)

    if len(first_line_clean) > 3 and len(first_line_clean) <= 80:
        return first_line_clean

    match = re.search(r'(?:Recipe|Dish|Name)[:\s]+([^\n]{3,60})', recipe_text, re.IGNORECASE)
    if match:
        return clean_dish_name(match.group(1))

    if len(first_line_clean) > 0:
        return first_line_clean

    return "unknown_recipe"


def repair_explanation_cache():
    print("=== Repairing explanation_cache.xlsx ===")
    wb = openpyxl.load_workbook(EXPLANATION_CACHE)
    ws = wb.active

    fixed_rows = []
    deleted_rows = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row_num = i + 2
        old_hash = row[0]
        old_recipe_text = row[1] or ""
        category = row[2]
        sugar_g = row[3]
        explanation = row[4]
        created_at = row[5]

        if not old_recipe_text or len(old_recipe_text.strip()) == 0:
            print(f"  Row {row_num}: SKIP (empty recipe_text) — deleting")
            ws.delete_rows(row_num)
            deleted_rows += 1
            continue

        normalized = clean_recipe_text(old_recipe_text)
        new_hash = hashlib.sha256(normalized.encode()).hexdigest()
        new_recipe_text = normalized[:100]

        ws.cell(row=row_num, column=1).value = new_hash
        ws.cell(row=row_num, column=2).value = new_recipe_text

        fix_type = "FIXED" if new_hash != old_hash else "clean"
        print(f"  Row {row_num}: {fix_type} | '{old_recipe_text[:40]}...'")
        print(f"    -> '{new_recipe_text[:40]}...'")
        fixed_rows.append(row_num)

    print(f"  Deleted {deleted_rows} empty rows")
    print(f"  Updated {len(fixed_rows)} rows with new hashes")

    wb.save(EXPLANATION_CACHE)
    print(f"  Saved to {EXPLANATION_CACHE}")


def repair_recipe_cache():
    print("\n=== Repairing recipe_cache.xlsx ===")
    wb = openpyxl.load_workbook(RECIPE_CACHE)

    if "recipes" not in wb.sheetnames:
        print("  No 'recipes' sheet found, skipping")
        return

    ws = wb["recipes"]

    rows_to_delete = []
    rows_to_update = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row_num = i + 2
        old_hash = row[0]
        old_dish_name = row[1] or ""
        recipe = row[2]
        created_at = row[3]

        dish_corrupted = '\n' in old_dish_name or len(old_dish_name) > 60

        if dish_corrupted and recipe:
            print(f"  Row {row_num}: CORRUPTED dish_name")
            print(f"    old: '{old_dish_name[:60]}...'")
            if len(old_dish_name) > 100:
                new_dish_name = extract_dish_name_from_recipe(recipe)
                print(f"    -> extracted: '{new_dish_name}'")
                rows_to_update.append((row_num, old_hash, new_dish_name.strip(), recipe, created_at))
            else:
                print(f"    -> UNRECOVERABLE, will delete row")
                rows_to_delete.append(row_num)
        else:
            cleaned = clean_dish_name(old_dish_name)
            new_hash = hashlib.sha256(cleaned.encode()).hexdigest()
            if new_hash != old_hash:
                print(f"  Row {row_num}: HASH_CHANGED | '{old_dish_name[:40]}' -> '{cleaned[:40]}'")
                print(f"    old_hash={str(old_hash)[:16]}... -> new_hash={new_hash[:16]}...")
                rows_to_update.append((row_num, old_hash, old_dish_name.strip(), recipe, created_at))
            else:
                print(f"  Row {row_num}: clean | '{old_dish_name[:40]}'")

    for row_num in reversed(rows_to_delete):
        print(f"  Deleting row {row_num}")
        ws.delete_rows(row_num)

    for row_num, old_hash, new_dish_name, recipe, created_at in rows_to_update:
        cleaned = clean_dish_name(new_dish_name)
        new_hash = hashlib.sha256(cleaned.encode()).hexdigest()
        ws.cell(row=row_num, column=1).value = new_hash
        ws.cell(row=row_num, column=2).value = new_dish_name[:100]
        print(f"  Updated row {row_num}: hash={new_hash[:16]}... | name='{new_dish_name[:40]}'")

    wb.save(RECIPE_CACHE)
    print(f"  Saved to {RECIPE_CACHE}")


if __name__ == "__main__":
    print("Starting cache repair...\n")
    repair_explanation_cache()
    repair_recipe_cache()
    print("\n=== Repair complete ===")