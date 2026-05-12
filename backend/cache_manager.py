import hashlib
import re
from datetime import datetime
from pathlib import Path

import openpyxl

EXPLANATION_CACHE = Path(__file__).parent / "explanation_cache.xlsx"
RECIPE_CACHE = Path(__file__).parent / "recipe_cache.xlsx"


def _clean_recipe_text(text):
    if not text:
        return ""
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'(?:^|(?<=\s))\*([^*/\s]+)\*(?=\s|$)', r'\1', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _clean_dish_name(text):
    if not text:
        return ""
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'(?:^|(?<=\s))\*([^*/\s]+)\*(?=\s|$)', r'\1', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


class ExplanationCache:
    def __init__(self):
        self.memory_cache = {}
        self._load_excel()

    def _hash_recipe(self, recipe: str) -> str:
        normalized = _clean_recipe_text(recipe)
        return hashlib.sha256(normalized.encode()).hexdigest()

    def _load_excel(self):
        if EXPLANATION_CACHE.exists():
            self.wb = openpyxl.load_workbook(EXPLANATION_CACHE)
            self.ws = self.wb.active
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[4]:
                    self.memory_cache[row[0]] = row[4]
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.append(
                ["recipe_hash", "recipe_text", "category", "sugar_g", "explanation", "created_at"]
            )
            self.wb.save(EXPLANATION_CACHE)

    def get(self, recipe: str, category: str):
        key = self._hash_recipe(recipe)
        return self.memory_cache.get(key)

    def set(self, recipe: str, category: str, sugar_g: float, explanation: str):
        key = self._hash_recipe(recipe)
        self.memory_cache[key] = explanation

        display_recipe = re.sub(r'\*\*([^*]+)\*\*', r'\1', recipe)
        display_recipe = re.sub(r'(?:^|(?<=\s))\*([^*/\s]+)\*(?=\s|$)', r'\1', display_recipe)

        self.ws.append(
            [
                key,
                display_recipe[:100],
                category,
                round(sugar_g, 2),
                explanation,
                datetime.now().strftime("%Y-%m-%d %H:%M"),
            ]
        )
        self.wb.save(EXPLANATION_CACHE)

    def get_cache_info(self):
        return {
            "total_entries": len(self.memory_cache),
            "cache_file": str(EXPLANATION_CACHE),
        }


class RecipeCache:
    def __init__(self):
        self.memory_cache = {}
        self._load_excel()

    def _hash_dish(self, dish_name: str) -> str:
        normalized = _clean_dish_name(dish_name)
        return hashlib.sha256(normalized.encode()).hexdigest()

    def _load_excel(self):
        if RECIPE_CACHE.exists():
            self.wb = openpyxl.load_workbook(RECIPE_CACHE)
        else:
            self.wb = openpyxl.Workbook()

        if "recipes" in self.wb.sheetnames:
            self.ws = self.wb["recipes"]
        else:
            self.ws = self.wb.create_sheet("recipes")
            self.ws.append(["dish_hash", "dish_name", "recipe", "created_at"])

        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[2]:
                self.memory_cache[row[0]] = row[2]

        self.wb.save(RECIPE_CACHE)

    def get(self, dish_name: str):
        normalized_key = _clean_dish_name(dish_name)
        key = hashlib.sha256(normalized_key.encode()).hexdigest()
        return self.memory_cache.get(key)

    def set(self, dish_name: str, recipe: str):
        if not dish_name or not dish_name.strip():
            print(f"[WARN] RecipeCache.set() rejected empty dish_name")
            return

        if '\n' in dish_name or len(dish_name) > 60:
            print(f"[WARN] RecipeCache.set() rejected dish_name with newline or too long: '{dish_name[:60]}...'")
            return

        normalized_key = _clean_dish_name(dish_name)
        key = hashlib.sha256(normalized_key.encode()).hexdigest()
        self.memory_cache[key] = recipe

        display_name = dish_name.strip()

        self.ws.append(
            [
                key,
                display_name[:100],
                recipe,
                datetime.now().strftime("%Y-%m-%d %H:%M"),
            ]
        )
        self.wb.save(RECIPE_CACHE)

    def get_cache_info(self):
        return {
            "total_entries": len(self.memory_cache),
            "cache_file": str(RECIPE_CACHE),
        }